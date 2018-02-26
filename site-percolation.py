'''
import openpyxl

wb = openpyxl.load_workbook('rough-1.xlsx')
#wb = openpyxl.load_workbook('rough.xlsx')

sheet = wb.get_active_sheet()

#for rows in sheet['A1':'E273']:
for rows in sheet['A1':'B5']:
        if rows == 'None':
            rows[0].value.delete
#    try:
	else:
#        print(rows[0].value + "\t" + rows[1].value + "\t" + rows[2].value + "\t" + rows[3].value + "\t" + rows[4].value)
	        print(rows[0].value + "\t" + rows[1].value)
    # TODO: remove rows with the word "None"
#    except:
#        if rows == "AP2M1":
#        if rows == "None":
#            rows[0].value.delete
#            continue
'''
import openpyxl
filename = 'rough-1.xlsx'
sheetname = 'Sheet1'

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name(sheetname)
for rows in sheet['A1':'E15250']:
  if rows[1].value == 'BCAS4' or rows[0].value == 'BCAS4' or rows[1].value == 'BCAS3' or rows[0].value == 'BCAS3':
# or rows[1].value == 'CUL3' or rows[0].value == 'CUL3' or rows[1].value == 'HNRNPA1' or rows[0].value == 'HNRNPA1' or rows[1].value == 'CPS5' or rows[0].value == 'CPS5' or rows[1].value == 'CAND1' or rows[0].value == 'CAND1' :

    rows[0].value = None 
    rows[1].value = None
    rows[2].value = None
    rows[3].value = None
    rows[4].value = None
  else:
    rows[0].value = str(rows[0].value) + ',' + str(rows[1].value) + ',' + str(rows[2].value) + ',' + str(rows[3].value) + ',' + str(rows[4].value)
#    rows[0].value = rows[0].value + ',' + rows[1].value + ',' + str(rows[2].value) 
    rows[1].value = None
    rows[2].value = None
    rows[3].value = None
    rows[4].value = None

wb.save(filename)
